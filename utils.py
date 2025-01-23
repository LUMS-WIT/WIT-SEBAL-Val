import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl
import os, re, csv, datetime, fnmatch
from osgeo import gdal
import osgeo.osr as osr
import metrics
from config import COMBINE_VALIDATIONS, INPUT_FOLDER, TEMPORAL_WIN

def save_metadata(metadata, filename):
    """
    Saves the given metadata list of dictionaries to an Excel file.
    
    Parameters:
    - metadata: List of dictionaries with metadata information.
    - filename: The name of the Excel file to save.
    """
    # Create a workbook and grab the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Define the header
    headers = ['gpi', 'latitude', 'longitude', 'start_date', 'end_date', 'count', 'overlaps']
    sheet.append(headers)
    
    # Add data from metadata to the sheet
    for data in metadata:
        row = [data.get(header, None) for header in headers]
        sheet.append(row)
    
    # Save the workbook to a file
    workbook.save(filename)
    print(f"Data saved to {filename}")

def save_to_plot(csv_dates, csv_values, raster_dates, raster_values, lat, lon, filename ):
    plt.figure(figsize=(12, 6))

    # Plot CSV data
    plt.plot(csv_dates, csv_values, marker='o', linestyle='-', color='b', label='WIT-SMS')

    # Plot Raster data
    plt.plot(raster_dates, raster_values, marker='x', linestyle='--', color='r', label='Sebal Soil Moisture')

    plt.title(f'Top Soil Moisture Time Series Comparison Latitude: {lat}, Longitude: {lon}')
    plt.xlabel('Date')
    plt.ylabel('Soil Moisture (m3/m3)')

    # Formatting the date on the x-axis
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.gca().xaxis.set_major_locator(mdates.AutoDateLocator())
    plt.gcf().autofmt_xdate()  # Auto formats the x-axis labels to fit them in the plot area

    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    # Save the plot to a file
    plt.savefig(filename, dpi=300)  # Save the figure to file
    print(f"Plot saved as {filename}")

def save_to_excel(data, filename, headers):
    """
    Saves given data to an Excel file with specified headers.
    
    Parameters:
    - data: List of tuples/lists containing the data to save.
    - filename: String for the filename to save the Excel document.
    - headers: List of strings for column headers in the Excel file.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Add headers to the first row
    sheet.append(headers)
    
    # Add data rows
    for row in data:
        sheet.append(row)
    
    # Save the workbook
    workbook.save(filename)
    print(f"Data saved to {filename}")

def remove_nan_entries(dates, values):
    """
    Removes entries with NaN values from date and value arrays.

    Parameters
    ----------
    dates : array-like
        Array of dates corresponding to the values.
    values : array-like
        Array of values associated with the dates.

    Returns
    -------
    tuple
        A tuple containing two arrays:
        - The cleaned array of dates with corresponding non-NaN values.
        - The cleaned array of values that are not NaN.
    """
    # Ensure the input arrays are numpy arrays
    dates = np.array(dates)
    values = np.array(values)

    # Create a mask to filter out NaN entries
    valid_indices = ~np.isnan(values)

    # Apply the mask to both the dates and values arrays
    cleaned_dates = dates[valid_indices]
    cleaned_values = values[valid_indices]

    return cleaned_dates.tolist(), cleaned_values.tolist()


class SoilMoistureData:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.data = []
        self.metadata = []
        self.total_files = 0

    def read_data(self):
        files = [f for f in os.listdir(self.folder_path) if f.endswith(".csv") and 'witsms_gpi' in f]
        self.total_files = len(files)
        print(f"Reading {self.total_files} files...")  # Debugging output

        for file in files:
            file_path = os.path.join(self.folder_path, file)
            gpi = re.search('gpi=(\d+)', file).group(1)
            lat = re.search('lat=([-+]?[0-9]*\.?[0-9]+)', file).group(1)
            lon = re.search('lon=([-+]?[0-9]*\.?[0-9]+)', file).group(1)
            timestamps = []
            soil_moistures = []

            with open(file_path, 'r') as csvfile:
                reader = csv.reader(csvfile)
                next(reader)  # Skip header
                for col in reader:
                    if col[1]:  # Ensuring the soil moisture value is not empty
                        try:
                            timestamp = datetime.datetime.strptime(col[0], "%Y-%m-%d %H:%M:%S")
                            soil_moisture = float(col[1]) / 100 # coverting to 0-1 scale
                            timestamps.append(timestamp)
                            soil_moistures.append(soil_moisture)
                        except ValueError as e:
                            print(f"Error parsing line in {file}: {e}")  # Error output

            # Ensure data is not empty before appending
            if timestamps:
                self.data.append((timestamps, soil_moistures))  # Storing timestamps and soil moisture values
                start_date = min(timestamps).strftime("%Y-%m-%d")
                end_date = max(timestamps).strftime("%Y-%m-%d")
                count_soil_moisture = len(soil_moistures)
                self.metadata.append({
                    'gpi': gpi,
                    'latitude': lat,
                    'longitude': lon,
                    'start_date': start_date,
                    'end_date': end_date,
                    'count': count_soil_moisture, 
                    'overlaps': 0
                })

    def print_metadata(self):
        headers = ['gpi', 'latitude', 'longitude', 'start_date', 'end_date', 'count', 'overlaps']
        print(",".join(headers))
        for entry in self.metadata:
            row = [entry[header] for header in headers]
            print(",".join(map(str, row)))
    
    def get_metadata(self):
        """
        Returns the metadata as a list of dictionaries.
        Each dictionary contains the metadata of one dataset.
        """
        return self.metadata

    def save_metadata_to_csv(self, filename='metadata.csv'):
        headers = ['gpi', 'latitude', 'longitude', 'start_date', 'end_date', 'count', 'overlaps']
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            for entry in self.metadata:
                writer.writerow(entry)
    
    def get_lat_lon_by_gpi(self, gpi):
        for item in self.metadata:
            if item['gpi'] == gpi:
                return (item['latitude'], item['longitude'])
        return None  # Return None if no matching GPI is found  

    def get_soil_moisture_by_location(self, lat=None, lon=None, gpi=None):
        if gpi:  # Get data by GPI if provided
            for (timestamps, soil_moistures), meta in zip(self.data, self.metadata):
                if meta['gpi'] == gpi:
                    return timestamps, soil_moistures
        elif lat and lon:  # Get data by latitude and longitude
            for (timestamps, soil_moistures), meta in zip(self.data, self.metadata):
                if meta['latitude'] == str(lat) and meta['longitude'] == str(lon):
                    return timestamps, soil_moistures
        return None, None  # Return None if no data is found

    def plot_data_gpi(self, gpi=None):
        # Plot data only for the specified GPI
        if gpi:
            found = False
            for (timestamps, soil_moistures), meta in zip(self.data, self.metadata):
                if meta['gpi'] == gpi:
                    plt.figure(figsize=(10, 6))
                    plt.plot(timestamps, soil_moistures, label=f"GPI {meta['gpi']} at ({meta['latitude']}, {meta['longitude']})")
                    plt.title(f"Soil Moisture Time Series for GPI {meta['gpi']} - Values: {meta['count']}")
                    plt.xlabel('Date')
                    plt.ylabel('Soil Moisture')
                    plt.legend()
                    plt.show()
                    found = True
                    break  # Stop after plotting the specified GPI
            if not found:
                raise ValueError(f'GPI {gpi} not available in metadata')
        else:
            # If no GPI is specified, plot all data
            for (timestamps, soil_moistures), meta in zip(self.data, self.metadata):
                plt.figure(figsize=(10, 6))
                plt.plot(timestamps, soil_moistures, label=f"GPI {meta['gpi']} at ({meta['latitude']}, {meta['longitude']})")
                plt.title(f"Soil Moisture Time Series for GPI {meta['gpi']} - Values: {meta['count']}")
                plt.xlabel('Date')
                plt.ylabel('Soil Moisture')
                plt.legend()
                plt.show()


class SebalSoilMoistureData:
    """
    Class for reading and processing raster data from a specified folder.
    Attributes:
        folder_path (str): Path to the folder containing raster files.
    """

    def __init__(self, folder_path):
        """
        Initializes the SebalData object.
        Args:
            folder_path (str): Path to the folder containing raster files.

        Raises:
            ValueError: If the folder path is invalid or does not exist.
        """

        if not os.path.exists(folder_path):
            raise ValueError(f"Invalid folder path: {folder_path}")
        self.folder_path = folder_path
        
        self.pattern = 'Top_soil_moisture'
        files = [f for f in os.listdir(self.folder_path) if f.endswith(".tif") and self.pattern in f]
        # taking a raster file for extent coverage check
        if len(files)>0:
            file = files[0]
            self.raster_file_path = os.path.join(self.folder_path, file)
        else:
            raise ValueError(f"Folder contains no raster files")   
        
        self.raster_dates = None 
        self.raster_values= None

    def get_raster_extent(self, file_path):
        """
        Returns the spatial extent of the given raster file.
        """
        dataset = gdal.Open(file_path)
        if not dataset:
            raise ValueError("Could not open the raster file.")

        # Get raster geometry info
        transform = dataset.GetGeoTransform()
        x_min = transform[0]
        y_max = transform[3]
        x_max = x_min + transform[1] * dataset.RasterXSize
        y_min = y_max + transform[5] * dataset.RasterYSize

        return (x_min, x_max, y_min, y_max)
    
    def is_within_extent(self, lat, lon, extent):
        """
        Check if the given latitude and longitude are within the raster extent.
        """
        x_min, x_max, y_min, y_max = extent
        return x_min <= lon <= x_max and y_min <= lat <= y_max
    
    def read_raster_value(self, file_path=None, lat=None, lon=None):
        """
        Reads the value at a specific latitude and longitude from a single raster file
        matching the given pattern within the folder.

        Args:
            lat (float): Latitude coordinate.
            lon (float): Longitude coordinate.
            pattern (str, optional): Pattern to match filenames (default: "Top_soil_moisture").

        Returns:
            float: Raster value at the specified coordinates, or np.nan if missing or out of bounds
        """
            # Open the TIF file
        dataset = gdal.Open(file_path)
        if not dataset:
            raise ValueError("File could not be opened.")
        
        # Fetch the geographic coordinate system and projection from the dataset
        old_cs = osr.SpatialReference()
        old_cs.ImportFromWkt(dataset.GetProjectionRef())
        
        # Create a new geographic coordinate system (WGS84)
        wgs84 = osr.SpatialReference()
        wgs84.ImportFromEPSG(4326)
        
        # Create a coordinate transformation from WGS84 to the dataset's projection
        transform = osr.CoordinateTransformation(wgs84, old_cs)
        
        # Transform the lat, lon coordinates to the dataset's projection
        x, y, z = transform.TransformPoint(lon, lat)
        
        # Get the affine transformation coefficients
        gt = dataset.GetGeoTransform()
        
        # Convert from geographic coordinates to raster pixel coordinates
        pixel_x = int((x - gt[0]) / gt[1])
        pixel_y = int((y - gt[3]) / gt[5])

        # Check if the pixel coordinates are within the raster bounds
        if (pixel_x < 0 or pixel_x >= dataset.RasterXSize or
            pixel_y < 0 or pixel_y >= dataset.RasterYSize):
            print(f"Requested pixel is out of bounds: ({pixel_x}, {pixel_y})")

        # Read the data value at the raster pixel coordinates
        band = dataset.GetRasterBand(1)
        value = band.ReadAsArray(pixel_x, pixel_y, 1, 1)
        
        # Close the dataset
        dataset = None
        
        # Handle missing data
        if value is None or np.isnan(value).all():
            return np.nan
        else:
            return value[0, 0]

    def read_data(self, lat=None, lon=None):
        """
        Reads values from all raster files matching the given pattern
        within the folder, at the specified latitude and longitude.

        Args:
            lat (float): Latitude coordinate.
            lon (float): Longitude coordinate.

        Returns:
            tuple(list, list) 
                - List of datetime objects (dates)
                - List of corresponding raster values (sorted by date)
        """
        values = []
        dates = []

        # Walk through the directory, including subdirectories
        for root, dirs, files in os.walk(self.folder_path):    
            # Iterate through all files in the folder
            for filename in files:
                if filename.endswith(".tif") and self.pattern in filename:
                    # Construct the full file path
                    file_path = os.path.join(root, filename)
                    # Extract date from filename
                    parts = filename.split('_')
                    date_str = f"{parts[-4]}_{parts[-3]}_{parts[-2]}"  # Assumes date is in the format {year}_{month}_{day}
                    date = datetime.datetime.strptime(date_str, '%Y_%m_%d')
                    # Read the raster value
                    value = self.read_raster_value(file_path, lat, lon)
                    values.append(value)
                    dates.append(date)
        
        # Combine dates and values into a list of tuples and sort by date
        _values = list(zip(dates, values))
        _values.sort(key=lambda x: x[0])  # Sorting by dates
        
        # Unpack dates and values from the sorted list
        raster_dates, raster_values = zip(*_values)

        return raster_dates, raster_values
    
    def get_data(self, lat, lon):
        """
        return the date and value data for the user
        """
        # TODO: needs to optimize, opening and closing raster twice
        extent = self.get_raster_extent(self.raster_file_path)
        if self.is_within_extent(lat, lon, extent):
            raster_dates, raster_values = self.read_data(lat, lon)       
            return raster_dates, raster_values
        else:
            return None, None
           
    def plot_data(self):
        
        if self.raster_dates:
            # Plotting the results of raster
            plt.figure(figsize=(10, 5))
            #plt.plot(dates, values, 'o')
            plt.plot(self.raster_dates, self.raster_values, marker='x', linestyle='--', color='r', label='Sebal Soil Moisture')
            plt.title(f'{self.pattern} Over Time')
            plt.xlabel('Date')
            plt.ylabel('Soil Moisture Value (m3/m3)')

            # Configure date format on x-axis
            # Set the x-ticks to match the dates of the data points
            plt.gca().set_xticks(self.raster_dates)
            plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.gcf().autofmt_xdate(rotation=90)  # Auto formats the x-axis labels to fit them in the plot area
            plt.gca().tick_params(axis='x', labelsize=8)  # Setting the font size to 8

            plt.legend()
            plt.tight_layout()
            plt.show()
        else:
            print('Please run read_data method first')


# STEP 2 functions

def extract_sm_data(xlsx_file_path,ref=1, test=2):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename=xlsx_file_path)
    sheet = workbook.active  # or workbook['SheetName'] if you know the sheet name

    # Initialize lists to store 'ref_sm' and 'test_sm' values
    test_sm_list = []   # x
    ref_sm_list = []    # y

    # Assuming the first row contains headers and data starts from the Reference row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Append the values from 'ref_sm' and 'test_sm' columns to their respective lists
        ref_sm_list.append(row[1])
        test_sm_list.append(row[2])

    # Convert lists to numpy arrays before returning
    ref_sm_array = np.array(ref_sm_list)
    test_sm_array = np.array(test_sm_list)

    return ref_sm_array, test_sm_array

def combine_val_files(root_dir, specific_folder_name):
    file_paths = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        if fnmatch.fnmatch(os.path.basename(dirpath), specific_folder_name):
            for file in filenames:
                if file.endswith('.xlsx') and 'witgpi' in file:
                    file_paths.append(os.path.join(dirpath, file))
    return file_paths

def validations(folder_path):
    # Initialize lists to compile data from all files
    wit_sm_ = []
    sebal_sm_ = []


    files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx") and 'witgpi' in f]
    total_files = len(files)
    print(f"Reading {total_files} files...")  # Debugging output

    observations = 0
    for file in files:
        wit_sm, sebal_sm = extract_sm_data(file)
                # x and y must have length at least 2.
        if len(wit_sm) > 1:
            observations += len(wit_sm)
            wit_sm_.extend(wit_sm)
            sebal_sm_.extend(sebal_sm)

    bias = metrics.bias(sebal_sm_, wit_sm_)
    mse, mse_corr, mse_bias, mse_var = metrics.mse(sebal_sm_, wit_sm_)
    ubrmsd = metrics.ubrmsd(sebal_sm_, wit_sm_)
    p_rho = metrics.pearson_r(sebal_sm_, wit_sm_)
    s_rho = metrics.spearman_r(sebal_sm_, wit_sm_)

    metrics_dict = {
        'bias': bias,
        'mse': mse,
        'ubrmsd': ubrmsd,
        'p_rho': p_rho,
        's_rho': s_rho
    }

    return metrics_dict, observations

def validations_gpi(folder_path, threshold= -0.5):
    
    # Initialize a dictionary to store lists of results for each metric
    metrics_dict = {
        'gpi': [],
        'bias': [],
        'mse': [],
        'ubrmsd': [],
        'p_rho': [],
        's_rho': []
    }

    if COMBINE_VALIDATIONS:
        specific_folder_name = f'*_{TEMPORAL_WIN}'
        files = combine_val_files(INPUT_FOLDER, specific_folder_name)
        total_files = len(files)
        print(f"Reading {total_files} files...")  # Debugging outpt
    
    else:    
        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx") and 'witgpi' in f]
        total_files = len(files)
        print(f"Reading {total_files} files...")  # Debugging output

    # defining pattern for extracting gpi
    pattern = r'witgpi_([^_]+)_'

    observations = 0
    for file in files:
        wit_sm, sebal_sm = extract_sm_data(file)

        gpi = re.search(pattern, file).group(1)
        # x and y must have length at least 2. otherwise p_rho wont work
        if len(wit_sm) > 1:

            bias = metrics.bias(sebal_sm, wit_sm)
            mse, mse_corr, mse_bias, mse_var = metrics.mse(sebal_sm, wit_sm)
            ubrmsd = metrics.ubrmsd(sebal_sm, wit_sm)
            p_rho = metrics.pearson_r(sebal_sm, wit_sm)
            s_rho = metrics.spearman_r(sebal_sm, wit_sm)

            # add a check to discard values with high negative corelation ( < 70%) and low points (<10)
            # Append all points with correlation greater than -0.5
            if p_rho >= threshold and s_rho >= threshold:
                observations += len(wit_sm)
                metrics_dict['gpi'].append(gpi)
                metrics_dict['bias'].append(bias)
                metrics_dict['mse'].append(mse)
                metrics_dict['ubrmsd'].append(ubrmsd)
                metrics_dict['p_rho'].append(p_rho)
                metrics_dict['s_rho'].append(s_rho)
            # Check for high negative correlation (< -0.7) with sufficient data points (> 10)
            elif (p_rho < threshold or s_rho < threshold) and len(wit_sm) > 10:
                observations += len(wit_sm)
                metrics_dict['gpi'].append(gpi)
                metrics_dict['bias'].append(bias)
                metrics_dict['mse'].append(mse)
                metrics_dict['ubrmsd'].append(ubrmsd)
                metrics_dict['p_rho'].append(p_rho)
                metrics_dict['s_rho'].append(s_rho)

    
    return metrics_dict, observations

def compute_statistics(data_dict):
    stats_results = {}
    for key, values in data_dict.items():
        if key == 'gpi':
            continue
        stats_results[key] = {}
        if key in ['p_rho', 's_rho']:
            stats_results[key]['mean'] = metrics.mean_r(values)
            stats_results[key]['median'] = np.nanmedian(values)
            stats_results[key]['IQR'] = metrics.calculate_iqr(values)
        else:
            stats_results[key]['mean'] = np.nanmean(values)
            stats_results[key]['median'] = np.nanmedian(values)
            stats_results[key]['IQR'] = metrics.calculate_iqr(values)
    return stats_results


def plot_box_and_whiskers(metrics_dict, filename= None, save= False):
    # Creating the plot
    fig, ax = plt.subplots()
    # Prepare data for plotting
    # Prepare data for plotting, excluding NaN values, gpi and CIs
    data_to_plot = [
        [value for value in metrics_dict[key] if not np.isnan(value)]
        for key in metrics_dict if key != 'gpi' and not key.endswith('_cl') and not key.endswith('_cu')
    ]
    # Create box plot
    bp = ax.boxplot(data_to_plot, showfliers=False) #, notch=True, patch_artist=True)

    # Annotating the median
    for i, line in enumerate(bp['medians']):
        x, y = line.get_xydata()[1]  # top of median line
        ax.annotate(f'{y:.3f}', xy=(x, y), xytext=(x, y+0.5), 
                    textcoords='offset points', ha='center', va='bottom')

    # # Annotating the mean
    # for i, data in enumerate(data_to_plot):
    #     mean_value = np.mean(data)  # Calculate mean
    #     ax.annotate(f'{mean_value:.3f}', xy=(i+1, mean_value), xytext=(0, 10),
    #                 textcoords='offset points', ha='center', va='bottom')
                
    # ax.set_title('Box and Whisker Plot for Metrics')
    ax.set_ylabel('Metric Values', fontsize = 15)
    ax.set_xlabel('Metrics', fontsize = 15)
    metrics_names = ['Bias \n (m³/m³)', 'mse \n (m³/m³)', 'Ubrmsd \n (m³/m³)', 'pearson r', 'spearman rho']
    # ax.set_xticklabels([key for key in metrics_dict.keys() if key != 'gpi' and not key.endswith('_cl') and not key.endswith('_cu')])
    ax.set_xticklabels([metric for metric in metrics_names], fontsize = 15)

    # Add horizontal grid lines
    ax.yaxis.grid(True, linestyle='-', which='major', color='lightgrey', alpha=0.5)
    plt.show()

    # Save the plot if requested
    if save:
        fig.savefig(filename)
        print(f'Plot saved as {filename}')

def plot_metric_with_ci(metrics_dict, metric, filename=None, save=False):
    """
    Plots a box-and-whiskers plot for a specified metric along with its confidence intervals.

    Parameters
    ----------
    metrics_dict : dict
        Dictionary containing metrics, including the specified metric and its confidence intervals.
    metric : str
        The metric to plot (e.g., 'ubrmsd', 'bias').
    filename : str, optional
        Filename to save the plot (default is None).
    save : bool, optional
        Whether to save the plot (default is False).
    """
    # Ensure required keys are present
    if metric not in metrics_dict or f"{metric}_cl" not in metrics_dict or f"{metric}_cu" not in metrics_dict:
        raise ValueError(f"Metrics dictionary must contain '{metric}', '{metric}_cl', and '{metric}_cu'.")

    # Prepare data for plotting
    data_to_plot = [
        [value for value in metrics_dict[f'{metric}_cl'] if not np.isnan(value)],
        [value for value in metrics_dict[metric] if not np.isnan(value)],
        [value for value in metrics_dict[f'{metric}_cu'] if not np.isnan(value)],
    ]

    # Create the plot
    fig, ax = plt.subplots()
    bp = ax.boxplot(data_to_plot, showfliers=False) #, notch=True, patch_artist=True)

    # Annotating the medians
    for i, line in enumerate(bp['medians']):
        x, y = line.get_xydata()[1]  # Top of median line
        ax.annotate(f'{y:.3f}', xy=(x, y), xytext=(0, 5),
                    textcoords='offset points', ha='right', va='bottom')

    # Adding CI annotations inside the graph
    annotation_text = (
        f"cl: Lower Confidence Interval\n"
        f"cu: Upper Confidence Interval"
    )
    ax.text(0.05, 0.95, annotation_text, transform=ax.transAxes,
            fontsize=15, color='black', ha='left', va='top')

    # Set title and labels
    ax.set_title(f'Plot for {metric.upper()} with Confidence Intervals', fontsize=15)
    if metric == 'bias':
        ax.set_ylabel('Bias (difference of means) in m³/m³', fontsize=15)
    else:
        ax.set_ylabel('Unbiased root-mean-square deviation in m³/m³', fontsize=15)
    ax.set_xticks([1, 2, 3])
    ax.set_xticklabels([f'{metric}_cl', f'{metric}', f'{metric}_cu'], fontsize=15)

    # Add grid lines
    ax.yaxis.grid(True, linestyle='-', which='major', color='lightgrey', alpha=0.5)

    # Show the plot
    plt.show()

    # Save the plot if requested
    if save and filename:
        fig.savefig(filename)
        print(f"Plot saved as {filename}")



def validations_gpi_adv(folder_path, threshold= -0.5, alpha=0.05):
    """
    Validates GPI metrics and calculates confidence intervals for ubrmsd.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing validation files.
    threshold : float, optional
        Threshold for correlation metrics (default is -0.5).
    alpha : float, optional
        Confidence level for ubrmsd CI (default is 0.05).

    Returns
    -------
    metrics_dict : dict
        Dictionary containing calculated metrics, including CI for ubrmsd.
    observations : int
        Total number of observations.
    """   
    # Initialize a dictionary to store lists of results for each metric
    # Initialize a dictionary to store lists of results for each metric
    metrics_dict = {
        'gpi': [],
        'bias_cl': [],
        'bias': [],
        'bias_cu': [],
        'mse': [],
        'ubrmsd_cl': [],  # Lower bound of CI
        'ubrmsd': [],
        'ubrmsd_cu': [],  # Upper bound of CI
        'p_rho': [],
        's_rho': []
    }

    if COMBINE_VALIDATIONS:
        specific_folder_name = f'*_{TEMPORAL_WIN}'
        files = combine_val_files(INPUT_FOLDER, specific_folder_name)
        total_files = len(files)
        print(f"Reading {total_files} files...")  # Debugging outpt
    
    else:    
        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx") and 'witgpi' in f]
        total_files = len(files)
        print(f"Reading {total_files} files...")  # Debugging output

    # defining pattern for extracting gpi
    pattern = r'witgpi_([^_]+)_'

    observations = 0
    for file in files:
        wit_sm, sebal_sm = extract_sm_data(file)

        gpi = re.search(pattern, file).group(1)
        # x and y must have length at least 2. otherwise p_rho wont work
        if len(wit_sm) > 1:

            bias = metrics.bias(sebal_sm, wit_sm)
            mse, mse_corr, mse_bias, mse_var = metrics.mse(sebal_sm, wit_sm)
            ubrmsd = metrics.ubrmsd(sebal_sm, wit_sm)
            p_rho = metrics.pearson_r(sebal_sm, wit_sm)
            s_rho = metrics.spearman_r(sebal_sm, wit_sm)

            # Compute confidence intervals
            ubrmsd_cl, ubrmsd_cu = metrics.ubrmsd_ci(sebal_sm, wit_sm, ubrmsd, alpha)
            bias_cl, bias_cu = metrics.bias_ci(sebal_sm, wit_sm, ubrmsd, alpha)

            # add a check to discard values with high negative corelation ( < 70%) and low points (<10)
            # Append all points with correlation greater than -0.5
            if p_rho >= threshold and s_rho >= threshold:
                observations += len(wit_sm)
                metrics_dict['gpi'].append(gpi)
                metrics_dict['bias_cl'].append(bias_cl)
                metrics_dict['bias'].append(bias)
                metrics_dict['bias_cu'].append(bias_cu)
                metrics_dict['mse'].append(mse)
                metrics_dict['ubrmsd_cl'].append(ubrmsd_cl)
                metrics_dict['ubrmsd'].append(ubrmsd)                
                metrics_dict['ubrmsd_cu'].append(ubrmsd_cu)
                metrics_dict['p_rho'].append(p_rho)
                metrics_dict['s_rho'].append(s_rho)
            # Check for high negative correlation (< -0.7) with sufficient data points (> 10)
            elif (p_rho < threshold or s_rho < threshold) and len(wit_sm) > 10:
                observations += len(wit_sm)
                metrics_dict['gpi'].append(gpi)
                metrics_dict['bias_cl'].append(bias_cl)
                metrics_dict['bias'].append(bias)
                metrics_dict['bias_cu'].append(bias_cu)
                metrics_dict['mse'].append(mse)
                metrics_dict['ubrmsd_cl'].append(ubrmsd_cl)
                metrics_dict['ubrmsd'].append(ubrmsd)
                metrics_dict['ubrmsd_cu'].append(ubrmsd_cu)
                metrics_dict['p_rho'].append(p_rho)
                metrics_dict['s_rho'].append(s_rho)

    
    return metrics_dict, observations